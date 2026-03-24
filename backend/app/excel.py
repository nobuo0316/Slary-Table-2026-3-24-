from io import BytesIO
from datetime import date
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Region_Code", "Region_Name", "Step_No", "Min_Daily_Wage", "Base_Monthly_Wage",
    "Step_Multiplier", "Monthly_Salary", "Semi_Month_1", "Semi_Month_2",
    "Currency", "Effective_Date", "Notes"
]

def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary_Table"

    # header
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # guidance row (optional)
    ws.freeze_panes = "A2"

    # set widths
    widths = [14, 22, 8, 14, 16, 16, 16, 14, 14, 10, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # add formulas for rows 2..(2+49*? ) - we keep formulas generic per row
    # Provide example lines for one region; user can copy down.
    # Base_Monthly_Wage = Min_Daily_Wage * 25
    # Monthly_Salary = Base_Monthly_Wage * Step_Multiplier
    # Semi_Month_1 = ROUNDDOWN(Monthly_Salary/2,2)
    # Semi_Month_2 = Monthly_Salary - Semi_Month_1
    example_rows = 10  # enough for guidance
    for r in range(2, 2 + example_rows):
        # formula columns: E, G, H, I
        ws[f"E{r}"] = f"=D{r}*25"
        ws[f"G{r}"] = f"=E{r}*F{r}"
        ws[f"H{r}"] = f"=ROUNDDOWN(G{r}/2,2)"
        ws[f"I{r}"] = f"=G{r}-H{r}"
        # defaults
        ws[f"J{r}"] = "PHP"

    # create two master sheets (optional but useful)
    ws2 = wb.create_sheet("Regions")
    ws2.append(["Region_Code", "Region_Name", "Min_Daily_Wage", "Effective_Date"])
    ws3 = wb.create_sheet("Steps")
    ws3.append(["Step_No", "Step_Multiplier"])
    for s in range(1, 50):
        ws3.append([s, ""])  # user fills multipliers

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def _cell_str(v) -> str:
    return "" if v is None else str(v).strip()

def parse_and_validate(file_bytes: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Returns:
      meta: {currency, effective_date}
      rows: list of row dicts matching HEADERS keys (normalized)
    Raises ValueError with aggregated error messages.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=False)
    if "Salary_Table" not in wb.sheetnames:
        raise ValueError("Sheet 'Salary_Table' not found.")

    ws = wb["Salary_Table"]

    # header validation
    header = [ _cell_str(ws.cell(row=1, column=c).value) for c in range(1, len(HEADERS)+1) ]
    if header != HEADERS:
        raise ValueError(f"Header mismatch. Expected: {HEADERS} / Got: {header}")

    errors = []
    rows = []
    seen = set()

    # read rows until empty Region_Code and Step_No
    for r in range(2, ws.max_row + 1):
        region_code = _cell_str(ws.cell(row=r, column=1).value)
        step_no_val = ws.cell(row=r, column=3).value

        if region_code == "" and (step_no_val is None or _cell_str(step_no_val) == ""):
            continue

        region_name = _cell_str(ws.cell(row=r, column=2).value)
        step_no = step_no_val
        min_daily = ws.cell(row=r, column=4).value
        base_monthly = ws.cell(row=r, column=5).value
        mult = ws.cell(row=r, column=6).value
        monthly = ws.cell(row=r, column=7).value
        semi1 = ws.cell(row=r, column=8).value
        semi2 = ws.cell(row=r, column=9).value
        currency = _cell_str(ws.cell(row=r, column=10).value) or "PHP"
        eff = ws.cell(row=r, column=11).value
        notes = _cell_str(ws.cell(row=r, column=12).value)

        # required checks
        if region_code == "":
            errors.append(f"Row {r}: Region_Code is required")
        if region_name == "":
            errors.append(f"Row {r}: Region_Name is required")
        if step_no is None:
            errors.append(f"Row {r}: Step_No is required")
        else:
            try:
                step_no = int(step_no)
            except:
                errors.append(f"Row {r}: Step_No must be integer")
        if min_daily is None:
            errors.append(f"Row {r}: Min_Daily_Wage is required")
        if mult is None:
            errors.append(f"Row {r}: Step_Multiplier is required")
        if eff is None:
            errors.append(f"Row {r}: Effective_Date is required")

        # value checks
        if currency != "PHP":
            errors.append(f"Row {r}: Currency must be PHP (got {currency})")

        # step range
        if isinstance(step_no, int) and not (1 <= step_no <= 49):
            errors.append(f"Row {r}: Step_No must be 1..49 (got {step_no})")

        # uniqueness
        if region_code and isinstance(step_no, int):
            key = (region_code, step_no)
            if key in seen:
                errors.append(f"Row {r}: Duplicate Region_Code + Step_No ({region_code}, {step_no})")
            seen.add(key)

        # numeric sanity
        def _to_float(x, name):
            if x is None:
                return None
            try:
                return float(x)
            except:
                errors.append(f"Row {r}: {name} must be number (got {x})")
                return None

        min_daily_f = _to_float(min_daily, "Min_Daily_Wage")
        mult_f = _to_float(mult, "Step_Multiplier")

        # derive / validate amounts
        if min_daily_f is not None:
            expected_base = round(min_daily_f * 25, 2)
            base_f = _to_float(base_monthly, "Base_Monthly_Wage") if base_monthly is not None else expected_base
            if base_f is not None and abs(base_f - expected_base) > 0.01:
                errors.append(f"Row {r}: Base_Monthly_Wage should be Min_Daily_Wage*25 (= {expected_base}), got {base_f}")

        if min_daily_f is not None and mult_f is not None:
            expected_monthly = round((min_daily_f * 25) * mult_f, 2)
            monthly_f = _to_float(monthly, "Monthly_Salary") if monthly is not None else expected_monthly
            if monthly_f is not None and monthly_f + 1e-9 < round(min_daily_f * 25, 2):
                errors.append(f"Row {r}: Monthly_Salary below minimum base (Min_Daily_Wage*25)")
            if monthly_f is not None and abs(monthly_f - expected_monthly) > 0.05:
                # allow small rounding differences
                errors.append(f"Row {r}: Monthly_Salary should be Base*Multiplier (= {expected_monthly}), got {monthly_f}")

            # semi-monthly checks
            semi1_f = _to_float(semi1, "Semi_Month_1")
            semi2_f = _to_float(semi2, "Semi_Month_2")
            if monthly_f is not None and semi1_f is not None and semi2_f is not None:
                if abs((semi1_f + semi2_f) - monthly_f) > 0.01:
                    errors.append(f"Row {r}: Semi_Month_1 + Semi_Month_2 must equal Monthly_Salary")
                # recommended rounding rule check (optional)
                expected_semi1 = float(int((monthly_f/2)*100))/100  # floor to 2 decimals
                if abs(semi1_f - expected_semi1) > 0.01:
                    errors.append(f"Row {r}: Semi_Month_1 should be ROUNDDOWN(Monthly/2,2) (= {expected_semi1})")

        # normalize effective date
        eff_date = None
        if eff is not None:
            if isinstance(eff, date):
                eff_date = eff
            else:
                # accept "YYYY-MM-DD"
                try:
                    y, m, d = map(int, str(eff).split("-"))
                    eff_date = date(y, m, d)
                except:
                    errors.append(f"Row {r}: Effective_Date must be date or YYYY-MM-DD (got {eff})")

        if not errors or True:
            # store raw; DB save will use validated computed numbers if needed
            rows.append({
                "Region_Code": region_code,
                "Region_Name": region_name,
                "Step_No": int(step_no) if isinstance(step_no, int) else step_no,
                "Min_Daily_Wage": float(min_daily_f) if min_daily_f is not None else None,
                "Base_Monthly_Wage": round(min_daily_f * 25, 2) if min_daily_f is not None else None,
                "Step_Multiplier": float(mult_f) if mult_f is not None else None,
                "Monthly_Salary": round((min_daily_f * 25) * mult_f, 2) if (min_daily_f is not None and mult_f is not None) else None,
                "Semi_Month_1": None,
                "Semi_Month_2": None,
                "Currency": currency,
                "Effective_Date": eff_date,
                "Notes": notes
            })

            # compute semi-monthly if possible
            if rows[-1]["Monthly_Salary"] is not None:
                monthly_salary = rows[-1]["Monthly_Salary"]
                semi1_calc = float(int((monthly_salary/2)*100))/100
                rows[-1]["Semi_Month_1"] = semi1_calc
                rows[-1]["Semi_Month_2"] = round(monthly_salary - semi1_calc, 2)

    if errors:
        # return all errors at once
        raise ValueError("\n".join(errors))

    # meta from first row
    currency_meta = rows[0]["Currency"] if rows else "PHP"
    eff_meta = rows[0]["Effective_Date"] if rows else None
    return {"currency": currency_meta, "effective_date": eff_meta}, rows

def export_version_to_excel(version_name: str, rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary_Table"
    ws.append(HEADERS)
    for r in rows:
        ws.append([
            r["region_code"], r["region_name"], r["step_no"], r["min_daily_wage"],
            r["base_monthly_wage"], r["step_multiplier"], r["monthly_salary"],
            r["semi_month_1"], r["semi_month_2"], r["currency"],
            r["effective_date"].isoformat() if r["effective_date"] else "", r.get("notes","")
        ])
    ws.freeze_panes = "A2"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
